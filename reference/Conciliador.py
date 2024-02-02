# -*- coding: utf-8 -*-
"""
Created on Sat Feb  4 17:53:41 2023

@author: Sebastián P. Pincheira
"""

# Importar paquetes
import pandas as pd
import numpy as np
import os
import dateutil.parser as dp
import datetime as dt
import tkinter as tk
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfilename
import tkinter.font as tkFont
from openpyxl import load_workbook
import tkinter.font as font

# Nombres de las variables
plata_M = "doc_sdo"
plata_C = "MONTO"
sucur_M = "suc_cod"
sucur_C = "NRO SUCURSAL"  # Esto va a mano!!!!!!!
# num_M     = se define en 'conciliar'
num_C = "N° DOCUMENTO"
fecha_M = "cofecha"
fecha_C = "FECHA"
detalle_M = "glosacon"
detalle_C = "DESCRIPCIÓN MOVIMIENTO"
prefacio_C = "Prefacio Cartola"  # Necesariamenta a mano.
prefacio_M = "Prefacio MAYORES"  # Necesariamente a mano.


def mesNum(n):
    if n in meses:
        return meses.index(n) + 1
    else:
        return


# GUI

window = tk.Tk()

#window.maxsize(619, 217)
window.minsize(619, 144)

# window.overrideredirect(True)

window.title("Conciliador")


# window.geometry("590x117")


def errorWindow(
    Cart, May, Empresa, Saldo_C, Saldo_M, Mes, Ano, Excel, wb, Format, num_M
):
    Condicion = [
        Mes != "Mes" and Ano != "Año",
        Empresa != "Empresa" and Saldo_C.isnumeric() and Saldo_M.isnumeric(),
        Format != 0,
        "MAYORES" in wb.sheetnames,
        "Cartola Santander" in wb.sheetnames,
        plata_M in May.columns,
        plata_C in Cart.columns,
        num_M in May.columns or "nrutfact" in May.columns,
        num_C in Cart.columns,
        fecha_M in May.columns,
        fecha_C in Cart.columns,
        detalle_M in May.columns,
        detalle_C in Cart.columns,
    ]

    if not False in Condicion:
        return True
    else:
        Outputs = [
            "Seleccione un mes y año válidos.",
            "Ingrese un nombre de empresa y saldos iniciales válidos.",
            "Seleccione una de las casillas 'numfact' o 'nrutfact'.",
            "No se encontró ninguna hoja con el nombre 'MAYORES' en el documento subido.",
            "No se encontró ninguna hoja con el nombre 'Cartola Santander' en el documento subido.",
            "No se encontró ninguna columna con el nombre '"
            + str(plata_M)
            + "' en MAYORES.",
            "No se encontró ninguna columna con el nombre '"
            + str(plata_C)
            + "' en Cartola Santander.",
            "No se encontró ninguna columna con el nombre '"
            + str(num_M)
            + "' o 'nrutfact' en MAYORES.",
            "No se encontró ninguna columna con el nombre '"
            + str(num_C)
            + "' en Cartola Santander.",
            "No se encontró ninguna columna con el nombre '"
            + str(fecha_M)
            + "' en MAYORES.",
            "No se encontró ninguna columna con el nombre '"
            + str(fecha_C)
            + "' en Cartola Santander.",
            "No se encontró ninguna columna con el nombre '"
            + str(detalle_M)
            + "' en MAYORES.",
            "No se encontró ninguna columna con el nombre '"
            + str(detalle_C)
            + "' en Cartola Santander.",
        ]
        for i in range(0, len(Condicion)):
            if Condicion[i] == False:
                say = Outputs[i]

                def on_enter_errorButton(e):
                    errorWindowButton["background"] = "#000D56"

                def on_leave_errorButton(e):
                    errorWindowButton["background"] = "#001693"

                errorWindow = tk.Toplevel(window)
                errorWindow.title("Error")
                errorWindow.config(bg=background)
                errorWindowFrame = tk.Frame(
                    errorWindow, bd=10, relief=tk.GROOVE, bg=background
                )
                errorWindowFrame.pack()
                errorWindowLabel = tk.Frame(errorWindowFrame, bd=5, bg=background)
                errorWindowExit = tk.Frame(errorWindowFrame, bd=10, bg=background)
                errorWindowLabel.pack()
                errorWindowExit.pack()
                errorWindow.title("Conciliador - Error")
                tk.Label(
                    errorWindowLabel,
                    text=say,
                    wraplength=200,
                    font=30,
                    bg=background,
                    fg="#DDDDDD",
                ).pack()
                errorWindowButton = tk.Button(
                    errorWindowExit,
                    text="OK",
                    font=40,
                    command=errorWindow.destroy,
                    bg="#001693",
                    fg="#FFFFFF",
                    bd=3,
                    width=10,
                    cursor="hand2",
                )
                errorWindowButton.pack()
                errorWindowButton.bind("<Enter>", on_enter_errorButton)
                errorWindowButton.bind("<Leave>", on_leave_errorButton)
                return False


def conciliar():
    # Variables GUI
    Excel = displayArchivo.get("1.0", tk.END).strip()
    Empresa = empresa.get("1.0", tk.END).strip()
    Saldo_C = saldo_C.get("1.0", tk.END).strip().replace(".", "")
    Saldo_M = saldo_M.get("1.0", tk.END).strip().replace(".", "")
    Mes = mesNum(mes.get())
    Ano = ano.get()
    Format = numfactVar.get() + nrutfactVar.get()
    if Format == 1:
        num_M = "numfact"
    else:
        num_M = "nrutfact"
    if not os.path.isfile(Excel):

        def on_enter_noFileButton(e):
            noFileButton["background"] = "#000D56"

        def on_leave_noFileButton(e):
            noFileButton["background"] = "#001693"

        noFile = tk.Toplevel(window)
        noFile.config(bg=background)
        noFileFrame = tk.Frame(noFile, bd=10, relief=tk.GROOVE, bg=background)
        noFileFrame.pack()
        noFileLabel = tk.Frame(noFileFrame, bd=5, bg=background)
        noFileExit = tk.Frame(noFileFrame, bd=10, bg=background)
        noFileLabel.pack()
        noFileExit.pack()
        noFile.title("Conciliador - Error")
        tk.Label(
            noFileLabel,
            text="No se pudo encontar el archivo: intente clickendo en 'Examinar...'.",
            wraplength=200,
            font=30,
            bg=background,
            fg="#DDDDDD",
        ).pack()
        noFileButton = tk.Button(
            noFileExit,
            text="OK",
            font=40,
            command=noFile.destroy,
            bg="#001693",
            fg="#FFFFFF",
            bd=3,
            width=10,
            cursor="hand2",
        )
        noFileButton.pack()
        noFileButton.bind("<Enter>", on_enter_noFileButton)
        noFileButton.bind("<Leave>", on_leave_noFileButton)
    else:
        wb = load_workbook(Excel, read_only=True)

        if "MAYORES" in wb.sheetnames:
            May = pd.read_excel(Excel, sheet_name="MAYORES".strip())
        else:
            May = pd.DataFrame({"": []})
        if "Cartola Santander" in wb.sheetnames:
            Cart = pd.read_excel(Excel, sheet_name="Cartola Santander".strip())
        else:
            Cart = pd.DataFrame({"": []})

        inputValido = errorWindow(
            Cart, May, Empresa, Saldo_C, Saldo_M, Mes, Ano, Excel, wb, Format, num_M
        )
        if inputValido == False:
            return
        else:
            # Importar cartola y mayores sin elementos NaN
            May = May.dropna(subset=[plata_M])
            Cart = Cart.dropna(subset=[plata_C])
            Cart[plata_C] = Cart[plata_C].astype(float)
            Saldo_C = Cart[plata_C].sum() + int(Saldo_C)
            Cart["Mes"] = (Cart[fecha_C].astype(str)).apply(lambda x: x[3:5])
            Cart = Cart[Cart["Mes"].astype(int) <= Mes]
            Cart[fecha_C] = (Cart[fecha_C]).astype(str).apply(dp.parse)

            May[plata_M] = May[plata_M].astype(float)
            May = May[May["cotipo"] != "A"]
            Saldo_M = May[plata_M].sum() + int(Saldo_M)
            May["Mes"] = (May[fecha_M].astype(str)).apply(lambda x: x[5:7])
            May = May[May["Mes"].astype(int) <= Mes]
            May[fecha_M] = (May[fecha_M]).astype(str).apply(dp.parse)
            Prefacio_C = pd.read_excel(Excel, sheet_name=prefacio_C).dropna(
                subset=[plata_C]
            )
            Prefacio_M = pd.read_excel(Excel, sheet_name=prefacio_M).dropna(
                subset=[plata_M]
            )
            Prefacio_M["cotipo"] = Prefacio_M[plata_M] * 0
            if Format == 2:
                Prefacio_M[num_M] = Prefacio_M["numfact"]
            if not sucur_C in Cart.columns:
                Cart[sucur_C] = Cart[plata_C] * 0
                May[sucur_M] = May[plata_M] * 0
                Prefacio_C[sucur_C] = Prefacio_C[plata_C] * 0
                Prefacio_M[sucur_M] = Prefacio_M[plata_M] * 0

            Cart = pd.concat(
                [
                    Prefacio_C[[fecha_C, detalle_C, num_C, sucur_C, plata_C]],
                    Cart[[fecha_C, detalle_C, num_C, sucur_C, plata_C]],
                ],
                axis=0,
            )
            May = pd.concat(
                [
                    Prefacio_M[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "cotipo"]],
                    May[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "cotipo"]],
                ],
                axis=0,
            )

            # Sacar los transacciones nulas de los Dataframes
            Cart = Cart[Cart[plata_C] != 0].reset_index(drop=True)
            May = May[May[plata_M] != 0].reset_index(drop=True)

            # Cración de columna "Key" que asocia a cada transacción, una única key
            Cart["Key"] = Cart[[plata_C, num_C, sucur_C]].apply(tuple, axis=1)
            May["Key"] = May[[plata_M, num_M, sucur_M]].apply(tuple, axis=1)

            # Diferenciar Keys repetidas agregando la cantidad de veces que ha salido el key antes en la última coordenada
            # A partir de ahora, los Key5 serán únicos y biyectarán las transacciones
            Zero_C = pd.DataFrame(0, index=np.arange(len(Cart)), columns=["Zero"])
            Zero_M = pd.DataFrame(0, index=np.arange(len(May)), columns=["Zero"])
            Zero_C["Count"] = Cart.groupby("Key").cumcount()
            Zero_M["Count"] = May.groupby("Key").cumcount()
            Cart["plata"] = np.abs(Cart[plata_C])
            May["plata"] = np.abs(May[plata_M])
            Cart["Key"] = Cart[["plata", num_C, sucur_C]].apply(tuple, axis=1)
            May["Key"] = May[["plata", num_M, sucur_M]].apply(tuple, axis=1)
            Zero_C["Key"] = Zero_C[["Count"]].apply(tuple, axis=1)
            Zero_M["Key"] = Zero_M[["Count"]].apply(tuple, axis=1)
            Zero_C["Key4"] = np.sign(Cart[plata_C])
            Zero_M["Key4"] = np.sign(May[plata_M])
            Zero_C["Key4"] = Zero_C[["Key4"]].apply(tuple, axis=1)
            Zero_M["Key4"] = Zero_M[["Key4"]].apply(tuple, axis=1)
            Cart["Key4"] = Cart["Key"] + Zero_C["Key"]
            May["Key4"] = May["Key"] + Zero_M["Key"]
            Cart["Key5"] = Cart["Key4"] + Zero_C["Key4"]
            May["Key5"] = May["Key4"] + Zero_M["Key4"]

            # Tomamos los elementos no paredos de cada lista
            NOCART = May[
                (~May["Key5"].isin(Cart["Key5"].array)) & (May["cotipo"] != "A")
            ].copy()
            NOCONT = Cart[(~Cart["Key5"].isin(May["Key5"].array))].copy()

            # DataFrames divididos por, repsectivamente, sin repeticiones de Key 4:
            # Abonos contabilizados no en cartola
            # Cargos contabilizados no en cartola
            # Abonos en cartola no contabilizados
            # Cargos en cartola no contabilizados
            NOCART["count"] = NOCART.groupby("Key4")["Key4"].transform("count")
            NOCONT["count"] = NOCONT.groupby("Key4")["Key4"].transform("count")

            NOCARTA = NOCART[(NOCART[plata_M] > 0) & (NOCART["count"] != 2)]
            NOCARTA = NOCARTA[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "count"]]

            NOCARTC = NOCART[(NOCART[plata_M] < 0) & (NOCART["count"] != 2)]
            NOCARTC = NOCARTC[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "count"]]

            NOCONTA = NOCONT[(NOCONT[plata_C] > 0) & (NOCONT["count"] != 2)]
            NOCONTA = NOCONTA[[fecha_C, detalle_C, num_C, sucur_C, plata_C, "count"]]

            NOCONTC = NOCONT[(NOCONT[plata_C] < 0) & (NOCONT["count"] != 2)]
            NOCONTC = NOCONTC[[fecha_C, detalle_C, num_C, sucur_C, plata_C, "count"]]

            # Los Key4 que se encontraron repetidos en el proceso anterior
            nocarta = NOCART[(NOCART[plata_M] > 0) & (NOCART["count"] != 1)]
            nocarta = nocarta[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "count"]]

            nocartc = NOCART[(NOCART[plata_M] < 0) & (NOCART["count"] != 1)]
            nocartc = nocartc[[fecha_M, detalle_M, num_M, sucur_M, plata_M, "count"]]

            noconta = NOCONT[(NOCONT[plata_C] > 0) & (NOCONT["count"] != 1)]
            noconta = noconta[[fecha_C, detalle_C, num_C, sucur_C, plata_C, "count"]]

            nocontc = NOCONT[(NOCONT[plata_C] < 0) & (NOCONT["count"] != 1)]
            nocontc = nocontc[[fecha_C, detalle_C, num_C, sucur_C, plata_C, "count"]]

            # Juntar las repetidas con las no repetidas
            noCartA = pd.concat([NOCARTA, nocarta], axis=0).reset_index(drop=True)
            noCartC = pd.concat([NOCARTC, nocartc], axis=0).reset_index(drop=True)
            noContA = pd.concat([NOCONTA, noconta], axis=0).reset_index(drop=True)
            noContC = pd.concat([NOCONTC, nocontc], axis=0).reset_index(drop=True)

            # Resultados
            SaldoConciliacion = (
                noCartA[plata_M].sum()
                + noCartC[plata_M].sum()
                + -noContA[plata_C].sum()
                + -noContC[plata_C].sum()
                + Saldo_C
            )

            Diferencia = SaldoConciliacion - Saldo_M

            file_name = (
                "Conciliacion_"
                + str(Empresa)
                + "_"
                + str(Mes)
                + "_"
                + str(Ano)
                + ".xlsx"
            )
            file_loc = file_name

            for i in [noCartA, noCartC]:
                i[fecha_M] = i[fecha_M].dt.strftime("%d-%m-%Y")
            for i in [noContA, noContC]:
                i[fecha_C] = i[fecha_C].dt.strftime("%d-%m-%Y")

            if os.path.exists(file_loc):
                os.remove(file_loc)

            # Cambiar nombres de columnas
            def Color():
                return ["background-color: red"]

            noCartA.rename(
                columns={
                    plata_M: "Monto",
                    detalle_M: "Detalle",
                    num_M: "N° Documento",
                    sucur_M: "Sucursal",
                    fecha_M: "Fecha",
                },
                inplace=True,
            )
            noCartC.rename(
                columns={
                    plata_M: "Monto",
                    detalle_M: "Detalle",
                    num_M: "N° Documento",
                    sucur_M: "Sucursal",
                    fecha_M: "Fecha",
                },
                inplace=True,
            )
            noContA.rename(
                columns={
                    plata_C: "Monto",
                    detalle_C: "Detalle",
                    num_C: "N° Documento",
                    sucur_C: "Sucursal",
                    fecha_C: "Fecha",
                },
                inplace=True,
            )
            noContC.rename(
                columns={
                    plata_C: "Monto",
                    detalle_C: "Detalle",
                    num_C: "N° Documento",
                    sucur_C: "Sucursal",
                    fecha_C: "Fecha",
                },
                inplace=True,
            )

            noContC.style.set_table_styles(
                [
                    {
                        "selector": "th",
                        "props": [("background-color", "black"), ("color", "cyan")],
                    }
                ]
            )

            with pd.ExcelWriter(file_loc) as writer:
                noCartAT = pd.DataFrame({"Abonos no en cartola": []})
                noCartAT.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1,
                    startcol=1,
                    index=False,
                )
                noCartA.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=2,
                    startcol=1,
                    index=False,
                )

                noCartCT = pd.DataFrame({"Cargos no en cartola": []})
                noCartCT.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + 3,
                    startcol=1,
                    index=False,
                )
                noCartC.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + 4,
                    startcol=1,
                    index=False,
                )

                noContAT = pd.DataFrame({"Abonos no contabilizados": []})
                noContAT.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + len(noCartC) + 6,
                    startcol=1,
                    index=False,
                )
                noContA.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + len(noCartC) + 7,
                    startcol=1,
                    index=False,
                )

                noContCT = pd.DataFrame({"Cargos no contabilizados": []})
                noContCT.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + len(noCartC) + len(noContA) + 9,
                    startcol=1,
                    index=False,
                )
                noContC.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1 + len(noCartA) + len(noCartC) + len(noContA) + 10,
                    startcol=1,
                    index=False,
                )

                dfSaldoConciliacion = pd.DataFrame(
                    {"Saldo Conciliación": [], int(SaldoConciliacion): []}
                )
                dfSaldoSegunMayor = pd.DataFrame(
                    {"Saldo Según Mayor": [], int(Saldo_M): []}
                )
                dfDiferencia = pd.DataFrame({"Diferencia": [], int(Diferencia): []})

                dfSaldoConciliacion.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1
                    + len(noCartA)
                    + len(noCartC)
                    + len(noContA)
                    + len(noContC)
                    + 12,
                    startcol=1,
                    index=False,
                )
                dfSaldoSegunMayor.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1
                    + len(noCartA)
                    + len(noCartC)
                    + len(noContA)
                    + len(noContC)
                    + len(dfSaldoConciliacion)
                    + 13,
                    startcol=1,
                    index=False,
                )
                dfDiferencia.to_excel(
                    writer,
                    sheet_name=str(mes.get()) + " - " + str(Ano)[2:4],
                    startrow=1
                    + len(noCartA)
                    + len(noCartC)
                    + len(noContA)
                    + len(noContC)
                    + len(dfSaldoConciliacion)
                    + len(dfSaldoSegunMayor)
                    + 14,
                    startcol=1,
                    index=False,
                )

            outputFrame.grid(column=1, row=4)
            output1.pack()
            output2.pack()
            output3.pack(anchor=tk.W)
            output1.config(
                text="La conciliación de "
                + str(empresa.get("1.0", tk.END).strip())
                + " del mes de "
                + str(mes.get()).lower()
                + " y año "
                + str(ano.get())
                + " se encuentra disponible en"
            )
            output2.config(text=str(os.path.abspath(file_loc)))
            output3.config(
                text="La diferencia de la conciliación es de "
                + str(int(Diferencia))
                + "."
            )


def browsefunc():
    outputFrame.grid_forget()
    output1.config(text="")
    output2.config(text="")
    output3.config(text="")
    filename = tk.filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
    displayArchivo.delete("1.0", tk.END)
    displayArchivo.insert(tk.END, filename)


def default_E(event):
    current_E = empresa.get("1.0", tk.END)
    if current_E == "Empresa\n":
        empresa.delete("1.0", tk.END)
    elif current_E == "\n":
        empresa.insert("1.0", "Empresa")


def default_C(event_C):
    current_C = saldo_C.get("1.0", tk.END)
    if current_C == "Saldo inicial Cartola\n":
        saldo_C.delete("1.0", tk.END)
    elif current_C == "\n":
        saldo_C.insert("1.0", "Saldo inicial Cartola")


def default_M(event_M):
    current_M = saldo_M.get("1.0", tk.END)
    if current_M == "Saldo inicial Mayores\n":
        saldo_M.delete("1.0", tk.END)
    elif current_M == "\n":
        saldo_M.insert("1.0", "Saldo inicial Mayores")


def default_A(event_A):
    current_A = displayArchivo.get("1.0", tk.END)
    if current_A == "Presione 'Examinar...' para seleccionar un archivo Excel\n":
        displayArchivo.delete("1.0", tk.END)
    elif current_A == "\n":
        displayArchivo.insert(
            "1.0", "Presione 'Examinar...' para seleccionar un archivo Excel"
        )


def on_enter_exit(e):
    Exit["background"] = "#6B0000"


def on_leave_exit(e):
    Exit["background"] = "#A00000"


def on_enter_conciliar(e):
    Conciliar["background"] = "#000D56"


def on_leave_conciliar(e):
    Conciliar["background"] = "#001693"


def on_enter_examinar(e):
    Examinar["background"] = "#DDDDDD"


def on_leave_examinar(e):
    Examinar["background"] = "#7A7A7A"


def on_enter_info(e):
    info["background"] = "#FCFF45"


def on_leave_info(e):
    info["background"] = "#DDDDDD"


def check1():
    nrutfactVar.set(0)


def check2():
    numfactVar.set(0)


background = "#000739"
window.config(bg=background)

mainFrame = tk.Frame(window, bd=10, relief=tk.GROOVE, bg=background)
mainFrame.pack()

examinFrame = tk.Frame(mainFrame, bd=4, bg=background)
examinFrame.grid(column=1, row=1)

entryFrame = tk.Frame(examinFrame, bd=6, relief=tk.GROOVE, bg=background)
entryFrame.pack(side=tk.LEFT)

typingFrame = tk.Frame(mainFrame, bd=5, bg=background)
typingFrame.grid(column=1, row=2, sticky=tk.W)

ajustesFrame = tk.Frame(mainFrame, bd=5, bg=background)
ajustesFrame.grid(column=1, row=3, sticky=tk.W)

finalFrame = tk.Frame(mainFrame, bd=4, bg=background)
finalFrame.grid(column=1, row=3, sticky=tk.E)

outputFrame = tk.Frame(
    mainFrame, bg=background, width=500, height=70, bd=5, relief=tk.SUNKEN
)

checkFrame = tk.Frame(ajustesFrame, bg=background)

displayArchivo = tk.Text(
    entryFrame,
    height=1,
    width=53,
    font=40,
    bg="#000000",
    fg="#FFFFFF",
    highlightthickness=1,
    insertbackground="#FFFFFF",
)
displayArchivo.pack()
displayArchivo.insert(
    tk.END, "Presione 'Examinar...' para seleccionar un archivo Excel"
)
displayArchivo.bind("<FocusIn>", default_A)
displayArchivo.bind("<FocusOut>", default_A)

Examinar = tk.Button(
    examinFrame,
    text="Examinar...",
    font=40,
    command=browsefunc,
    bd=4,
    bg="#7A7A7A",
    activebackground="#DDDDDD",
    cursor="based_arrow_up",
)
Examinar.pack(side=tk.RIGHT)
Examinar.bind("<Enter>", on_enter_examinar)
Examinar.bind("<Leave>", on_leave_examinar)

dropDownFont = tkFont.Font(size=12)
dropDownFontItems = tkFont.Font(size=10)

now = dt.datetime.now()

meses = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

anos = []
anoActual = int(now.strftime("%Y"))
for ano in range(min(2021, anoActual), max(2021, anoActual) + 1):
    anos.append(ano)
    ano += 1

ano = tk.StringVar()

ano.set("Año")

mes = tk.StringVar()

mes.set("Mes")


drop_mes = tk.OptionMenu(ajustesFrame, mes, *meses)
drop_mes["highlightthickness"] = 0
drop_mes["bg"] = "#959595"
drop_mes["menu"].config(bg="#959595")
drop_mes["activebackground"] = "#DDDDDD"
drop_mes["cursor"] = "hand2"
drop_mes.config(font=dropDownFont, width=9)

menuMes = ajustesFrame.nametowidget(drop_mes.menuname)
menuMes.config(font=dropDownFontItems)

drop_ano = tk.OptionMenu(ajustesFrame, ano, *anos)
drop_ano["highlightthickness"] = 0
drop_ano["bg"] = "#959595"
drop_ano["menu"].config(bg="#959595")
drop_ano["activebackground"] = "#DDDDDD"
drop_ano["cursor"] = "hand2"
drop_ano.config(font=dropDownFont, width=4)

menuAno = ajustesFrame.nametowidget(drop_ano.menuname)
menuAno.config(font=dropDownFontItems)

drop_mes.pack(side=tk.LEFT)

drop_ano.pack(side=tk.LEFT)

checkFrame.pack(side=tk.LEFT)

numfactVar = tk.IntVar()

numfactCheck = tk.Checkbutton(
    checkFrame,
    variable=numfactVar,
    text="numfact",
    onvalue=1,
    offvalue=0,
    bg=background,
    font=30,
    fg="#DDDDDD",
    selectcolor="#000000",
    activebackground=background,
    activeforeground="#DDDDDD",
    cursor="hand2",
    command=check1,
)
numfactCheck.pack(side=tk.LEFT)

nrutfactVar = tk.IntVar()

nrutfactCheck = tk.Checkbutton(
    checkFrame,
    variable=nrutfactVar,
    text="nrutfact",
    onvalue=2,
    offvalue=0,
    bg=background,
    font=30,
    fg="#DDDDDD",
    selectcolor="#000000",
    activebackground=background,
    activeforeground="#DDDDDD",
    cursor="hand2",
    command=check2,
)
nrutfactCheck.pack(side=tk.LEFT)

empresa = tk.Text(
    typingFrame,
    height=1,
    width=18,
    font=35,
    bd=3,
    highlightthickness=1,
    highlightbackground="#000000",
)
empresa.insert(tk.END, "Empresa")
empresa.pack(side=tk.LEFT)
empresa.bind("<FocusIn>", default_E)
empresa.bind("<FocusOut>", default_E)

saldo_C = tk.Text(
    typingFrame,
    height=1,
    width=18,
    font=35,
    bd=3,
    highlightthickness=1,
    highlightbackground="#000000",
)
saldo_C.insert(tk.END, "Saldo inicial Cartola")
saldo_C.pack(side=tk.LEFT)
saldo_C.bind("<FocusIn>", default_C)
saldo_C.bind("<FocusOut>", default_C)

saldo_M = tk.Text(
    typingFrame,
    height=1,
    width=18,
    font=35,
    bd=3,
    highlightthickness=1,
    highlightbackground="#000000",
)
saldo_M.insert(tk.END, "Saldo inicial Mayores")
saldo_M.pack(side=tk.LEFT)
saldo_M.bind("<FocusIn>", default_M)
saldo_M.bind("<FocusOut>", default_M)

Exit = tk.Button(
    finalFrame,
    text="Salir",
    font=40,
    command=window.destroy,
    bg="#A00000",
    fg="#FFFFFF",
    bd=3,
    activebackground="#DDDDDD",
    cursor="hand2",
)
Exit.pack(side=tk.RIGHT)
Exit.bind("<Enter>", on_enter_exit)
Exit.bind("<Leave>", on_leave_exit)

Conciliar = tk.Button(
    finalFrame,
    text="Conciliar",
    font=40,
    command=conciliar,
    bg="#001693",
    fg="#FFFFFF",
    bd=3,
    activebackground="#DDDDDD",
    cursor="hand2",
)
Conciliar.pack(side=tk.RIGHT)
Conciliar.bind("<Enter>", on_enter_conciliar)
Conciliar.bind("<Leave>", on_leave_conciliar)


def question():
    def on_enter_errorButton(e):
        errorWindowButton["background"] = "#000D56"

    def on_leave_errorButton(e):
        errorWindowButton["background"] = "#001693"

    say = (
        "Los elementos necesarios en el documento Excel son los siguientes (y deben tener exactamente el mismo nombre): "
        + "Una hoja con el nombre 'Cartola Santander', una hoja con el nombre 'MAYORES', una hoja con el nombre 'Prefacio Cartola' "
        + "y una hoja con el nombre 'Prefacio MAYORES'. La 'Cartola Santander' corresponde a la información del banco y tiene que tener las columnas: "
        + "'MONTO', 'NRO SUCURSAL' (si se desea utilizar, si no, no es necesario incluir), 'N° DOCUMENTO', 'FECHA' y 'DESCRIPCIÓN MOVIMIENTO'. El 'Prefacio Cartola' contiene a todos los elementos de la cartola de años anteriores "
        + "y contiene, por lo menos, las mismas columnas que tiene 'Cartola Santander'. 'MAYORES' constituye la información contabilziada "
        + "y tiene que tener las columnas: 'doc_sdo', 'suc_cod', 'numfact' (y 'nrutfact'), 'cofecha' y 'glosacon'. La hoja 'Prefacio MAYORES' consiste en los elementos contabilizados "
        + "que no se encuentran en el banco de años anteriores. Su información consiste en, por lo menos, las mismas columnas que 'MAYORES' exceptuando, quizás, 'cotipo'. Si no se encuentra una columna "
        + "con el nombre 'NRO SUCURSAL' en 'Cartola Santander', se ignorarán todas las columnas de ese tipo en todas las hojas."
    )
    errorWindow = tk.Toplevel(window)
    errorWindow.config(bg=background)
    errorWindow.title("Conciliador - Información")
    errorWindowFrame = tk.Frame(
        errorWindow, bd=10, relief=tk.GROOVE, bg=background, width=300
    )
    exitFrame = tk.Frame(errorWindowFrame, bd=10, bg=background, width=300)
    datosFrame = tk.Frame(exitFrame, bd=10, bg=background, width=100)
    errorWindowFrame.pack()
    errorWindowLabel = tk.Frame(errorWindowFrame, bd=5, bg=background, width=300)
    datosLabel1 = tk.Label(
        datosFrame, text="Sebastián P. Pincheira", bg=background, fg="#DDDDDD", bd=0
    )
    datosLabel1_ = tk.Label(
        datosFrame,
        text="Estudiante Ing. Matemática/Creador del Programa",
        bg=background,
        fg="#DDDDDD",
        bd=0,
    )
    datosLabel2 = tk.Label(
        datosFrame, text="+56 9 8918 6914 (WhatsApp)", bg=background, fg="#DDDDDD", bd=0
    )
    datosLabel3 = tk.Label(
        datosFrame, text="spincheiral@gmail.com", bg=background, fg="#DDDDDD", bd=0
    )
    ESPACIOXD = tk.Label(exitFrame, text="jshczxbdf", bg=background, fg=background)
    # errorWindowExit   = tk.Frame(exitFrame, bd=10, bg = background)
    errorWindowLabel.pack()
    datosFrame.grid(column=1, row=1)
    ESPACIOXD.grid(column=2, row=1)
    datosLabel1.pack(anchor=tk.W)
    datosLabel1_.pack(anchor=tk.W)
    datosLabel2.pack(anchor=tk.W)
    datosLabel3.pack(anchor=tk.W)
    # errorWindowExit.pack()
    tk.Label(
        errorWindowLabel, text=say, wraplength=800, font=30, bg=background, fg="#DDDDDD"
    ).pack()

    exitFrame.pack(anchor=tk.W)
    errorWindowButton = tk.Button(
        exitFrame,
        text="OK",
        font=40,
        command=errorWindow.destroy,
        bg="#001693",
        fg="#FFFFFF",
        bd=3,
        width=10,
        cursor="hand2",
    )
    errorWindowButton.grid(column=3, row=1)
    errorWindowButton.bind("<Enter>", on_enter_errorButton)
    errorWindowButton.bind("<Leave>", on_leave_errorButton)


info = tk.Button(
    finalFrame, text="?", width=1, command=question, anchor=tk.SW, cursor="hand2"
)
info["font"] = font.Font(size=8)
info.bind("<Enter>", on_enter_info)
info.bind("<Leave>", on_leave_info)

info.pack()

output1 = tk.Label(outputFrame, width=81, bg=background, fg="#DDDDDD")

output2 = tk.Label(outputFrame, width=81, bg=background, fg="#DDDDDD")

output3 = tk.Label(outputFrame, width=81, bg=background, fg="#DDDDDD")

window.mainloop()
